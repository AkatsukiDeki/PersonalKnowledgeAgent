"""Parser factory — dispatches file bytes to the correct format parser."""

import json
import logging
from typing import Dict, Any

logger = logging.getLogger(__name__)

# Mapping of file extensions to their MIME-like categories
SUPPORTED_EXTENSIONS: Dict[str, str] = {
    ".md": "text",
    ".txt": "text",
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".csv": "csv",
    ".json": "json",
    ".png": "image",
    ".jpg": "image",
    ".jpeg": "image",
    ".webp": "image",
    ".heic": "image",
}


def get_file_extension(filename: str) -> str:
    """Return normalised lowercase extension including the dot."""
    import os
    _, ext = os.path.splitext(filename.lower())
    return ext


def is_supported(filename: str) -> bool:
    """Check whether we have a parser for the given filename."""
    return get_file_extension(filename) in SUPPORTED_EXTENSIONS


def parse_file(filename: str, file_bytes: bytes) -> tuple[str, str, Dict[str, Any]]:
    """Parse a file and return (normalised_text, file_type, metadata).

    Parameters
    ----------
    filename : str
        Original filename (used to pick the parser).
    file_bytes : bytes
        Raw content of the uploaded file.

    Returns
    -------
    tuple[str, str, dict]
        ``(normalised_text, file_type, metadata)``
    """
    ext = get_file_extension(filename)
    if ext not in SUPPORTED_EXTENSIONS:
        raise ValueError(f"Unsupported file extension: {ext}")

    metadata: Dict[str, Any] = {
        "original_filename": filename,
        "file_size_bytes": len(file_bytes),
    }

    if ext in (".md", ".txt"):
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("cp1251", errors="replace")
        metadata["encoding"] = "utf-8"
        return text, ext.lstrip("."), metadata

    if ext == ".pdf":
        from .pdf_parser import parse_pdf
        text = parse_pdf(file_bytes)
        # Count pages for metadata
        from pypdf import PdfReader
        from io import BytesIO
        reader = PdfReader(BytesIO(file_bytes))
        metadata["page_count"] = len(reader.pages)
        return text, "pdf", metadata

    if ext == ".docx":
        from .docx_parser import parse_docx
        text = parse_docx(file_bytes)
        return text, "docx", metadata

    if ext == ".xlsx":
        from .tabular_parser import parse_xlsx
        text = parse_xlsx(file_bytes)
        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True)
        metadata["sheet_count"] = len(wb.sheetnames)
        metadata["sheet_names"] = wb.sheetnames
        wb.close()
        return text, "xlsx", metadata

    if ext == ".csv":
        from .tabular_parser import parse_csv
        text = parse_csv(file_bytes)
        row_count = text.count("\n") - 1  # minus header + separator
        metadata["row_count_approx"] = max(0, row_count)
        return text, "csv", metadata

    if ext == ".json":
        try:
            raw = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            raw = file_bytes.decode("cp1251", errors="replace")
        data = json.loads(raw)
        text = json.dumps(data, ensure_ascii=False, indent=2)
        metadata["json_keys"] = list(data.keys()) if isinstance(data, dict) else None
        return text, "json", metadata
        
    if ext in (".png", ".jpg", ".jpeg", ".webp", ".heic"):
        from .image_parser import parse_image
        text = parse_image(file_bytes, filename)
        return text, "image", metadata

    raise ValueError(f"No parser available for extension: {ext}")
